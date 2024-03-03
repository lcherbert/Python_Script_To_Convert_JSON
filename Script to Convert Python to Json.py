from openpyxl import load_workbook
from json import dumps
 
# Load Excel workbook
wb = load_workbook("D:\MSC DATA SCIENCE AND ARTIFICIAL INTELLIGENCE\Module 2 - Software Development in Practice\End of Module Assignment (Group Project)\Dictionary_To_Convert.xlsx")
 
# Choose a specific sheet
sheet = wb["Dictionary"]
 
# Find the number of rows and columns in the sheet
rows = sheet.max_row
columns = sheet.max_column
 
# List to store all rows as dictionaries
lst = []
 
# Iterate over rows and columns to extract data
for i in range(1, rows):
    row = {}
    for j in range(1, columns):
        column_name = sheet.cell(row=1, column=j)
        row_data = sheet.cell(row=i+1, column=j)
 
        row.update(
            {
                column_name.value: row_data.value
            }
        )
    lst.append(row)
 
# Convert extracted data into JSON format
json_dict = dumps(lst)
 
# Print the JSON data
print(json_dict)

json_file_path = "D:\MSC DATA SCIENCE AND ARTIFICIAL INTELLIGENCE\Module 2 - Software Development in Practice\End of Module Assignment (Group Project)\Json_Dictionary.txt"

with open (json_file_path, "w") as file:
    file.write(json_dict)

print(f"The Json Dictionary has been saved to the .txt file. ")
